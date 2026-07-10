"""
Submission Module
"""

from __future__ import annotations
from src.config import CONFIG
from pathlib import Path

import openpyxl
import pandas as pd


FRAMEWORK_SECTIONS = [
    "Variables Used",
    "Profitability Equation",
    "Prediction Logic",
    "Variable Selection Logic",
    "Coefficient/Weight Derivation",
    "Feature Transformations",
    "Business Logic",
    "Assumptions",
    "Validation Approach",
    "Additional Notes (Optional)",
]


def create_submission(
    df: pd.DataFrame,
    score: pd.Series,
    path=CONFIG.submission_path,
    framework: dict[str, str] | None = None,
    template_path: Path | str = CONFIG.submission_template_path,
) -> pd.DataFrame:

    path = Path(path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    submission = pd.DataFrame(
        {
            "ID": df["ID"],
            "Prediction": score,
        }
    )

    if path.suffix.lower() == ".xlsx":
        write_submission_workbook(
            submission=submission,
            path=path,
            framework=framework or {},
            template_path=template_path,
        )
    else:
        ranked_submission = submission.copy()
        ranked_submission["Prediction"] = (
            score.rank(
                ascending=False,
                method="first",
            ).astype(int)
        )
        ranked_submission.to_csv(
            path,
            index=False,
        )
        submission = ranked_submission

    print(f"\nSubmission saved to:\n{path}")

    return submission


def write_submission_workbook(
    submission: pd.DataFrame,
    path: Path,
    framework: dict[str, str],
    template_path: Path | str = CONFIG.submission_template_path,
) -> None:
    workbook = openpyxl.load_workbook(Path(template_path))

    predictions_sheet = workbook["Predictions"]

    for row_index, row in enumerate(
        submission.itertuples(index=False),
        start=2,
    ):
        predictions_sheet.cell(row=row_index, column=1, value=int(row.ID))
        predictions_sheet.cell(row=row_index, column=2, value=float(row.Prediction))

    framework_sheet = workbook["Profitability Framework"]

    for row_index in range(2, framework_sheet.max_row + 1):
        section = framework_sheet.cell(row=row_index, column=1).value

        if section in FRAMEWORK_SECTIONS:
            framework_sheet.cell(
                row=row_index,
                column=2,
                value=framework.get(section, ""),
            )

    workbook.save(path)

def save_raw_scores(
    df: pd.DataFrame,
    path=CONFIG.raw_score_path,
) -> pd.DataFrame:

    path = Path(path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    scores = pd.DataFrame()

    scores["ID"] = df["ID"]

    scores["Prediction"] = df["BEST_Score"]

    scores.to_csv(
        path,
        index=False,
    )

    print(f"\nRaw scores saved to:\n{path}")

    return scores
